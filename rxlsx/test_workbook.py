from decimal import Decimal
from rxlsx import Workbook


def test_sheet_title():
    wb = Workbook()
    ws1 = wb.create_sheet()
    assert ws1.title == "Sheet"

    ws2 = wb.create_sheet("Testerion")
    assert ws2.title == "Testerion"

    ws3 = wb.create_sheet(title="Keyword title")
    assert ws3.title == "Keyword title"


def test_sheet_index():
    wb = Workbook()
    ws1 = wb.create_sheet()

    assert wb.index(ws1) == 0

    ws2 = wb.create_sheet()
    assert wb.index(ws2) == 1

    ws3 = wb.create_sheet(index=0)
    assert wb.index(ws3) == 0

    assert wb.index(ws1) == 1
    assert wb.index(ws2) == 2


def test_sheet_append():
    wb = Workbook()
    ws = wb.create_sheet()
    assert ws.max_row_idx == 0
    assert ws.max_col_idx == 0
    ws.append([1, 2])
    assert ws.max_row_idx == 1
    assert ws.max_col_idx == 2


def test_workbook_gc():
    wb = Workbook()
    print(f"Inital refcount: {wb._refcount()}")

    ws1 = wb.create_sheet()
    ws2 = wb.create_sheet()

    print(f"Two sheets: {wb._refcount()}")

    wb.remove_sheet(ws1)
    wb.remove_sheet(ws2)

    print(f"Sheets removed: {wb._refcount()}")

    del ws1
    del ws2

    print(f"Sheets deleted: {wb._refcount()}")
    del wb


def test_save():
    import os
    from pathlib import Path

    wb = Workbook()
    ws = wb.create_sheet()
    ROW = [
        1,
        2,
        3,
        4,
        5,
        6,
        7,
        8,
        9,
        10,
        True,
        False,
        0.1,
        Decimal("0.2"),
        "foo",
        "bar",
        "baz",
        "qux",
        "fred",
        "thud",
        "=SUM(1+1)",
        "=SUM(A1+B1)",
        "string with space",
        "string with escapes < &",
        "" # empty string
    ]
    ws.append(ROW)
    ws.append(list(reversed(ROW)))

    string_path = "./a.zip"  # os.devnull
    wb.save(string_path)

    string_path = "./b.xlsx"
    file = open(string_path, "wb+")
    wb.save(file)

    from openpyxl import load_workbook

    wb = load_workbook(file)
    print(wb.sheetnames)
    ws = wb.active
    print(ws["A1"].value)

    # pathlib_path = Path(string_path)
    # wb.save(pathlib_path)
